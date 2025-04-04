# ============================== Python Libraries ==============================
import warnings
warnings.filterwarnings("ignore")
from flask import Flask, render_template, request
import yfinance as yf
import pandas as pd
import numpy as np
from statsmodels.tsa.arima.model import ARIMA
import datetime
from openpyxl import load_workbook, Workbook
from static_data import assets_names, index, mutual_funds, fd_rates, stepwise_model_order

# ============================== Global Variables ==============================
trained_models = {}
last_update_date = None

# ============================== General Functions ==============================
def read_excel(i, filename):
    data = []
    wb = load_workbook(filename)
    for row in wb[wb.sheetnames[i]].iter_rows(values_only=True):
        data.append(row)
    return pd.DataFrame(data[1:], columns=data[0])['Close']

def get_top_banks(age_group):
    column = "Return Rate for Adults (%)" if age_group == "adult" else "Return Rate for Senior Citizens (%)"
    top_3 = fd_rates.nlargest(3, column)[["Bank Name", column]]
    return top_3.to_dict(orient="records")  # Convert to JSON-friendly format

def generate_closing_prices():
    np.random.seed()
    close_prices = np.cumsum(np.random.randn(1000) * 2 + 100)
    return pd.DataFrame({'Close': close_prices})

def get_data(stock_name):
    ticker_data = yf.Ticker(stock_name)
    historical_data = ticker_data.history(period='max')
    if len(historical_data) > 0:
        df = pd.DataFrame(historical_data)
        return df
    else:
        print(f"No data for {stock_name}")

def train_arima_model(asset_data, stepwise_order=(5, 2, 0)):
    model = ARIMA(asset_data, order=stepwise_order)
    model_fit = model.fit()
    return model_fit

def update_models():
    global trained_models, last_update_date
    last_update_date = datetime.date.today()
    
    # assets = {
    #     "Index": {ind: generate_closing_prices() for ind in index},
    #     "Mutual_Funds": {fund: generate_closing_prices() for fund in mutual_funds},
    #     "Gold_Bond": {'gold_bond': generate_closing_prices()}
    # }

    assets = {
        "Index": {ind: read_excel(i, 'Index_Data.xlsx') for i, ind in enumerate(index)},
        "Mutual_Funds": {fund: read_excel(i, 'Mutual_Funds_Data.xlsx') for i, fund in enumerate(mutual_funds)},
        "Gold_Bond": {'gold_bond': read_excel(0, 'Gold_Bond_Data.xlsx')}
    }

    trained_models = {
        "Index": {name: train_arima_model(df, stepwise_model_order["Index"][name]) for name, df in assets["Index"].items()},
        "Mutual_Funds": {name: train_arima_model(df, stepwise_model_order["Mutual_Funds"][name]) for name, df in assets["Mutual_Funds"].items()},
        "Gold_Bond": {name: train_arima_model(df, stepwise_model_order["Gold_Bond"][name]) for name, df in assets["Gold_Bond"].items()}
    }



app = Flask(__name__)

with app.app_context():
    update_models()

@app.before_request
def check_for_update():
    if datetime.date.today() != last_update_date:
        update_models()

@app.route("/")
def index():
    return render_template("index.html", show_result=False)

@app.route("/about")
def about():
    return render_template("about.html")

@app.post('/run_model')
def run_model():
    risk_level = 'low'
    risk_level = request.form.get('risklevel')
    time_period = int(request.form.get('time_period'))
    investment_amount = int(request.form.get('investment_amount'))
    age = request.form.get('age')

    try:
        best_asset_return = {
            "risk_level": risk_level,
            "time_period": time_period,
            "investment_amount": investment_amount,
            "age": age,
            'Top_3_FD_Banks': get_top_banks(age),
            "Index": {
                "best_asset": None,
                "best_return": float('-inf') # negative infinity, so any new value will be greater than this
            },
            "Mutual_Funds": {
                "best_asset": None,
                "best_return": float('-inf')
            },
            "Gold_Bond": {
                "best_asset": None,
                "best_return": float('-inf')
            },
        }

        for asset_type, model_dict in trained_models.items():
            if isinstance(model_dict, dict):
                for name, model in model_dict.items():
                    forecast = model.forecast(steps=time_period).iloc[-1]
                    last_price = model.model.endog[-1]
                    percentage_change = ((forecast - last_price) / last_price) * 100

                    if percentage_change > best_asset_return[asset_type]['best_return']:
                        best_asset_return[asset_type]['best_return'] = percentage_change
                        best_asset_return[asset_type]['best_asset'] = assets_names[name]

    except Exception as e:
        print("An unexpected error occurred:")
        print(e)


    return render_template('index.html', show_result=True, result=best_asset_return)

if __name__=="__main__":
    app.run(debug=True, port=8000)